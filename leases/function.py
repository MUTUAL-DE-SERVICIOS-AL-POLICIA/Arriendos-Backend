from django.shortcuts import render
from django.http import HttpResponse
from .models import Rental
from .serializer import RentalsSerializer
from weasyprint import HTML, CSS
from django.template.loader import render_to_string
import os
from openpyxl import Workbook
from django.template.loader import render_to_string
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill
from datetime import datetime
from plans.models import Plan
from decimal import Decimal
def Make_Delivery_Form(request, rental_id, product):
    serializer_class = RentalsSerializer
    rental = rental_id
    rental = Rental.objects.get(pk = rental_id)
    serializer = serializer_class(rental)
    rental = serializer.data
    contract_number = rental.get("contract_number")
    customer = rental.get("customer")
    contacts = customer.get("contacts")
    customers = customer.get("contacts")

    customer = customers[0]
    name = customer.get("name")
    nit = customer.get("ci_nit")
    nup = customer.get("nup")

    selected_products_data = rental.get("selected_products")
    selected_products=[]
    for selected_product_data in selected_products_data:
        if selected_product_data["id"] == product:
            selected_product = {
                'id': selected_product_data["id"],
                'start_time': selected_product_data["start_time"],
                'product':selected_product_data["product"],
                'event_type':selected_product_data["event_type"]
            }
            selected_products.append(selected_product)

    user = request.user
    today = datetime.now()
    date = today.strftime("%d/%m/%y")
    ruta_archivo_html = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'entrega_y_recepcion_de_ambientes.html')
    ruta_logo = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logo.jpg')
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="entrega_y_recepcion_de_ambientes.pdf"'
    if os.path.exists(ruta_archivo_html):

        with open(ruta_archivo_html, 'r', encoding='utf-8') as archivo_html:
                html_content = archivo_html.read()

    html_string = render_to_string('entrega_y_recepcion_de_ambientes.html', {
        'name': name,
        'nit': nit,
        'selected_products':selected_products,
        'warranty': rental.get("initial_total"),
        'contacts': contacts,
        'contract_number': contract_number,
        'nup': nup,
        'product':product,
        'date': date,
        'user': user,
        'logo': 'file://' + ruta_logo
        })

    HTML(string=html_string).write_pdf(response, stylesheets=[CSS(
        string='@page { margin-left: 2cm; margin-right: 1cm; margin-top: 0.2cm; margin-bottom: 2cm; }'
        )])
    return response

def Make_Overtime_Form(request, rental_id, product):
    serializer_class = RentalsSerializer
    rental = rental_id
    rental = Rental.objects.get(pk = rental_id)
    serializer = serializer_class(rental)
    rental = serializer.data
    contract_number = rental.get("contract_number")
    customer = rental.get("customer")
    contacts = customer.get("contacts")
    institution_name = customer.get("institution_name", None)
    institution_nit = customer.get("nit", None)
    customers = customer.get("contacts")
    customer = customers[0]
    name = customer.get("name")
    nit = customer.get("ci_nit")
    nup = customer.get("nup")
    if institution_name is None:
        detail_name = name
        deatil_nit = nit
    else:
        detail_name = institution_name
        deatil_nit = institution_nit
    selected_products_data = rental.get("selected_products")
    selected_products=[]
    for selected_product_data in selected_products_data:
        if selected_product_data["id"] == product:
            additional_hour_applieds_data = selected_product_data["additional_hour_applieds"]
            additional_hour_applieds = []
            for additional_hour_applied in additional_hour_applieds_data:
                payment_date = additional_hour_applied["created_at"]
                try:
                    payment_date = datetime.strptime(payment_date, "%Y-%m-%dT%H:%M:%S.%f%z")
                except ValueError:
                    payment_date = datetime.strptime(payment_date, "%Y-%m-%dT%H:%M:%S%z")
                payment_date = payment_date.strftime("%d-%m-%Y %I:%M:%S %p")
                additional_hour_applied["rate_data"] = additional_hour_applied["total"] / additional_hour_applied["number"]
                additional_hour_applied["payment_date"] = payment_date
            selected_product = {
                'id': selected_product_data["id"],
                'start_time': selected_product_data["start_time"],
                'product':selected_product_data["product"],
                'additional_hour_applieds': additional_hour_applieds_data
            }
            selected_products.append(selected_product)
    user = request.user
    today = datetime.now()
    date = today.strftime("%d/%m/%y")
    ruta_archivo_html = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'horas_extra.html')
    ruta_logo = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logo.jpg')
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="horas_extra.pdf"'
    if os.path.exists(ruta_archivo_html):

        with open(ruta_archivo_html, 'r', encoding='utf-8') as archivo_html:
                html_content = archivo_html.read()

    html_string = render_to_string('horas_extra.html', {
        'name': name,
        'nit': nit,
        'institution_name': institution_name,
        'institution_nit': institution_nit,
        'detail_name': detail_name,
        'detail_nit': deatil_nit,
        'selected_products':selected_products,
        'contacts': contacts,
        'contract_number': contract_number,
        'nup': nup,
        'date': date,
        'user': user,
        'logo': 'file://' + ruta_logo
        })

    HTML(string=html_string).write_pdf(response, stylesheets=[CSS(
        string='@page { margin-left: 2cm; margin-right: 1cm; margin-top: 0.2cm; margin-bottom: 2cm; }'
        )])
    return response

def Make_Rental_Report(request, start_date, end_date, state):
    if start_date is None and end_date is None:
        rentals = Rental.objects.filter(state_id=state).distinct()
    else:
        rentals = Rental.objects.filter(
                    selected_product__start_time__range=(start_date, end_date),
                    state_id=state
                ).distinct()

    serializer_class = RentalsSerializer
    serializer = serializer_class(rentals, many=True)
    wb = Workbook()
    ws = wb.active

    switch_headers = {
        1: ["EDIFICIO", "SALON", "EVENTO", "FECHA DEL EVENTO", "CONTACTO", "NIT/CI", "TELEFONO", "INSTITUCION", "PLAN"],
        2: ["EDIFICIO", "SALON", "EVENTO", "FECHA DEL EVENTO", "ALQUILER", "CONTACTO", "NIT/CI", "TELEFONO","INSTITUCION", "PLAN"],
        3: ["EDIFICIO", "SALON", "EVENTO", "FECHA DEL EVENTO", "ALQUILER", "PAGO DEL ALQUILER", "FECHA DEL PAGO", "GARANTIA DEL ALQUILER", "FECHA DE PAGO DE LA GARANTIA", "CONTACTO", "NIT/CI", "TELEFONO","INSTITUCION", "PLAN", "HORAS EXTRA", "PRECIO TOTAL HORAS EXTRA"],
        4: ["EDIFICIO", "SALON", "EVENTO", "FECHA DEL EVENTO", "ALQUILER", "PAGO DEL ALQUILER", "FECHA DEL PAGO", "GARANTIA DEL ALQUILER", "FECHA DE PAGO DE LA GARANTIA", "FECHA SOLICITUD DEVOLUCION", "FECHA DEVOLUCION", "MONTO RETORNADO", "CONTACTO", "NIT/CI", "TELEFONO","INSTITUCION", "PLAN", "HORAS EXTRA", "PRECIO TOTAL HORAS EXTRA"],
        5: ["EDIFICIO", "SALON", "EVENTO", "FECHA DEL EVENTO", "ALQUILER", "PAGO DEL ALQUILER", "FECHA DEL PAGO", "GARANTIA DEL ALQUILER", "FECHA DE PAGO DE LA GARANTIA", "FECHA SOLICITUD DEVOLUCION", "FECHA DEVOLUCION", "MONTO RETORNADO", "CONTACTO", "NIT/CI", "TELEFONO","INSTITUCION", "PLAN", "HORAS EXTRA", "PRECIO TOTAL HORAS EXTRA"]
    }

    headers = switch_headers.get(state, [])
    ws.append(headers)
    rentals = serializer.data

    for rentals in serializer.data:
        customers = rentals.get("customer")
        institution_name = customers["institution_name"]
        if institution_name is None:
            institution_name = "Ninguno"
        contacts = customers.get("contacts")
        if contacts:
            first_contact = contacts[0]
            name = first_contact.get("name")
            ci_nit = first_contact.get("ci_nit")
            phone = first_contact.get("phone")

        plan = rentals["plan"]
        if plan:
            plan_name_obj = Plan.objects.get(pk=plan)
            plan_name = plan_name_obj.plan_name
        else:
            plan_name = "Ninguno"

        selected_products = rentals.get("selected_products")
        contract_number = rentals["contract_number"]

        warranty_movements = rentals.get("warranty_movements")
        payments = rentals.get("payments")
        last_record = payments[-1] if payments else {}
        payments_last = last_record.get("amount_paid", "")
        payable_mount = last_record.get("payable_mount", "")
        voucher_number_payment = last_record.get("voucher_number", "")
        payment_date = last_record.get("created_at", "")
        payment_total = 0
        for payment in payments:
            payment_total = payment_total + Decimal(payment["amount_paid"])

        if payment_date:
            try:
                payment_date = datetime.strptime(payment_date, "%Y-%m-%dT%H:%M:%S.%f%z")
            except ValueError:
                payment_date = datetime.strptime(payment_date, "%Y-%m-%dT%H:%M:%S%z")
            payment_date = payment_date.strftime("%d-%m-%Y %I:%M:%S %p")

        warranty_return_request = rentals["warranty_return_request"]
        if warranty_return_request:
            try:
                warranty_return_request = datetime.strptime(warranty_return_request, "%Y-%m-%dT%H:%M:%S.%f%z")
            except ValueError:
                warranty_return_request = datetime.strptime(warranty_return_request, "%Y-%m-%dT%H:%M:%S%z")

            warranty_return_request = warranty_return_request.strftime("%d-%m-%Y %I:%M:%S %p")
        else:
            warranty_return_request = "No se realizo la solicitud"

        warranty_returned = rentals["warranty_returned"]
        if warranty_returned:
            try:
                warranty_returned = datetime.strptime(warranty_returned, "%Y-%m-%dT%H:%M:%S.%f%z")
            except ValueError:
                warranty_returned = datetime.strptime(warranty_returned, "%Y-%m-%dT%H:%M:%S%z")

            warranty_returned = warranty_returned.strftime("%d-%m-%Y %I:%M:%S %p")

        else:
            warranty_returned = "No se retornó la garantía"
        warranty_movements = rentals.get("warranty_movements", [])

        get_warranty = warranty_movements[0] if warranty_movements else {}
        get_warranty_income = get_warranty.get("income", "")
        get_warranty_voucher_number = get_warranty.get("voucher_number", "")
        get_warranty_created_at = get_warranty.get("created_at", "")
        if get_warranty_created_at:
            try:
                get_warranty_datetime = datetime.strptime(get_warranty_created_at, "%Y-%m-%dT%H:%M:%S.%f%z")
            except ValueError:
                get_warranty_datetime = datetime.strptime(get_warranty_created_at, "%Y-%m-%dT%H:%M:%S%z")
            get_warranty_created_at = get_warranty_datetime.strftime("%d-%m-%Y %I:%M:%S %p")

        warranty_total = 0
        for warranty_movement in warranty_movements:
            warranty_total = warranty_total + Decimal(warranty_movement["income"])
        get_warranty_last_record = warranty_movements[-1] if warranty_movements else {}
        get_warranty_returned = get_warranty_last_record.get("returned", "")

        current_contract_number = None

        for selected_product in selected_products:
            room = selected_product.get("product", {})
            room = room.get("room", {})
            room_name = room.get("name", "")
            property = room.get("property")
            property_name = property["name"]
            events = selected_product.get("event_type")
            event_name = events.get("name")
            room_capacity = room.get("capacity", "")

            hours_additional_applied = selected_product.get('additional_hour_applieds', '')
            get_hour_additional_applied = hours_additional_applied[-1] if hours_additional_applied else {}
            number_hours_applied = get_hour_additional_applied.get("number", "")
            voucher_number_hours_applied = get_hour_additional_applied.get("voucher_number", "")
            
            total_number_hours_applied = 0
            total_hours_applied = 0
            for hours_additional in hours_additional_applied:
                total_number_hours_applied = total_number_hours_applied + hours_additional["number"]
                total_hours_applied = total_hours_applied + hours_additional["total"]
            payment_total_data = 0
            warranty_total_data = None
            warranty_returned_data = None
            if plan:
                if current_contract_number is not None and current_contract_number == contract_number:
                    payment_total_data = payment_total
                    warranty_total_data = warranty_total
                    warranty_returned_data = get_warranty_returned
                else:
                    payment_total_data = None
                    warranty_total_data = None
                    warranty_returned_data = None
                current_contract_number = contract_number
            else:
                payment_total_data = payment_total
                warranty_total_data = warranty_total
                warranty_returned_data = get_warranty_returned

            switch_data = {
                1: [property_name,
                    room_name,
                    event_name,
                    selected_product["start_time"],
                    name,
                    ci_nit,
                    phone,
                    institution_name,
                    plan_name,
                    ],
                2: [property_name,
                    room_name,
                    event_name,
                    selected_product["start_time"],
                    contract_number,
                    name,
                    ci_nit,
                    phone,
                    institution_name,
                    plan_name
                    ],
                3: [
                    property_name,
                    room_name,
                    event_name,
                    selected_product["start_time"],
                    contract_number,
                    payment_total_data,
                    payment_date,
                    warranty_total_data,
                    get_warranty_created_at,
                    name,
                    ci_nit,
                    phone,
                    institution_name,
                    plan_name,
                    total_number_hours_applied,
                    total_hours_applied,
                ],
                4: [
                    property_name,
                    room_name,
                    event_name,
                    selected_product["start_time"],
                    contract_number,
                    payment_total_data,
                    payment_date,
                    warranty_total_data,
                    get_warranty_created_at,
                    warranty_return_request,
                    warranty_returned,
                    warranty_returned_data,
                    name,
                    ci_nit,
                    phone,
                    institution_name,
                    plan_name,
                    total_number_hours_applied,
                    total_hours_applied
                ],
                5: [
                    property_name,
                    room_name,
                    event_name,
                    selected_product["start_time"],
                    contract_number,
                    payment_total_data,
                    payment_date,
                    warranty_total_data,
                    get_warranty_created_at,
                    warranty_return_request,
                    warranty_returned,
                    warranty_returned_data,
                    name,
                    ci_nit,
                    phone,
                    institution_name,
                    plan_name,
                    total_number_hours_applied,
                    total_hours_applied
                ]
            }
            data = switch_data.get(state, [])
            ws.append(data)

    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
        if idx % 2 == 1:
            for cell in row:
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        else:
            for cell in row:
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = Border(
                    left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"),
                )
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_de_alquileres.xlsx'
    wb.save(response)
    return response
